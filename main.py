# ============================================================
# main.py — Linen Tracker API
# ============================================================

from fastapi import FastAPI, HTTPException
from sqlalchemy.orm import Session
from database import SessionLocal, Towel, Event, DeletedTag, init_db
from pydantic import BaseModel
from typing import Optional
import datetime
import asyncio
from contextlib import asynccontextmanager
from fastapi.middleware.cors import CORSMiddleware
import os
from fastapi.security.api_key import APIKeyHeader
from fastapi import Security
from fastapi import FastAPI, HTTPException, Depends
from fastapi.security import HTTPBasic, HTTPBasicCredentials
import secrets
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# ============================================================
# HTTP BASIC AUTH
# ============================================================
security = HTTPBasic()

def verify_staff(credentials: HTTPBasicCredentials = Depends(security)):
    correct_user = secrets.compare_digest(
        credentials.username, os.getenv("STAFF_USER", "staff")
    )
    correct_pass = secrets.compare_digest(
        credentials.password, os.getenv("STAFF_PASS", "linen2026")
    )
    if not (correct_user and correct_pass):
        raise HTTPException(
            status_code=401,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Basic"},
        )
    return credentials.username

# ============================================================
# API KEY AUTH
# ============================================================
API_KEY = os.getenv("API_KEY", "dev-key")
api_key_header = APIKeyHeader(name="X-API-Key", auto_error=False)

def verify_key(key: str = Security(api_key_header)):
    if key != API_KEY:
        raise HTTPException(status_code=403, detail="Invalid or missing API key")

# ============================================================
# WASH CYCLE LIMITS
# ============================================================
MAX_WASHES = 200
WASH_WARNING  = 150
WASH_CRITICAL = 190

# ============================================================
# EMAIL
# ============================================================
def send_email(subject: str, html_body: str):
    gmail_user    = os.getenv("GMAIL_USER")
    gmail_pass    = os.getenv("GMAIL_PASS")
    manager_email = os.getenv("MANAGER_EMAIL")
    if not all([gmail_user, gmail_pass, manager_email]):
        print("Email not sent — GMAIL_USER, GMAIL_PASS or MANAGER_EMAIL not set")
        return
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"]    = gmail_user
        msg["To"]      = manager_email
        msg.attach(MIMEText(html_body, "html"))
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(gmail_user, gmail_pass)
            server.sendmail(gmail_user, manager_email, msg.as_string())
        print(f"Email sent to {manager_email}: {subject}")
    except Exception as e:
        print(f"Email failed: {e}")

# ============================================================
# BACKGROUND JOB — AUTO MISSING DETECTION
# ============================================================
async def auto_mark_missing():
    while True:
        await asyncio.sleep(3600)
        db = SessionLocal()
        try:
            cutoff = datetime.datetime.utcnow() - datetime.timedelta(hours=24)
            missing_towels = db.query(Towel).filter(
                Towel.status == "in_use",
                Towel.dispatched_at < cutoff
            ).all()
            for towel in missing_towels:
                towel.status = "missing"
                event = Event(
                    tag_id=towel.tag_id,
                    event_type="MISSING",
                    location=towel.last_location,
                    created_at=datetime.datetime.utcnow()
                )
                db.add(event)
            db.commit()
            print(f"Missing check done — {len(missing_towels)} towel(s) flagged")
        finally:
            db.close()

# ============================================================
# DAILY REPORT
# ============================================================
async def daily_report():
    while True:
        now = datetime.datetime.utcnow()
        next_8am = now.replace(hour=8, minute=0, second=0, microsecond=0)
        if now >= next_8am:
            next_8am += datetime.timedelta(days=1)
        seconds_until_8am = (next_8am - now).total_seconds()
        print(f"Daily report scheduled in {seconds_until_8am/3600:.1f} hours")
        await asyncio.sleep(seconds_until_8am)

        db = SessionLocal()
        try:
            total          = db.query(Towel).count()
            registered     = db.query(Towel).filter(Towel.status == "registered").count()
            in_use         = db.query(Towel).filter(Towel.status == "in_use").count()
            in_laundry     = db.query(Towel).filter(Towel.status == "in_laundry").count()
            missing        = db.query(Towel).filter(Towel.status == "missing").count()
            missing_towels = db.query(Towel).filter(Towel.status == "missing").all()
            recent_deleted = db.query(DeletedTag).filter(
                DeletedTag.deleted_at >= datetime.datetime.utcnow() - datetime.timedelta(hours=24)
            ).all()
        finally:
            db.close()

        is_urgent = missing >= 5
        subject = (
            f"🚨 URGENT — {missing} Towels Missing | Linen Report"
            if is_urgent else
            f"📊 Daily Linen Report — {datetime.datetime.utcnow().strftime('%d %b %Y')}"
        )

        if missing_towels:
            missing_rows = "".join([
                f"<tr><td style='padding:8px'>{t.tag_id}</td>"
                f"<td style='padding:8px'>{t.towel_type or '—'}</td>"
                f"<td style='padding:8px'>{t.last_location or '—'}</td></tr>"
                for t in missing_towels
            ])
            missing_section = f"<h3>Missing Towels</h3><table>{missing_rows}</table>"
        else:
            missing_section = "<p style='color:#166534'>✓ No missing towels today</p>"

        if recent_deleted:
            deleted_rows = "".join([
                f"<tr><td style='padding:8px'>{d.tag_id}</td>"
                f"<td style='padding:8px'>{d.towel_type or '—'}</td>"
                f"<td style='padding:8px'>{d.total_washes}</td>"
                f"<td style='padding:8px'>{d.reason or '—'}</td></tr>"
                for d in recent_deleted
            ])
            deleted_section = f"<h3>Deleted Today</h3><table>{deleted_rows}</table>"
        else:
            deleted_section = ""

        urgent_banner = (
            f"<div style='background:#dc2626;color:white;padding:12px'>⚠ ALERT: {missing} towels missing over 24 hours.</div>"
            if is_urgent else ""
        )

        html_body = (
            "<div style='font-family:Arial,sans-serif;max-width:600px;margin:0 auto'>"
            "<div style='background:#1e3a5f;padding:20px'>"
            "<h1 style='color:white;margin:0'>Smart Linen Tracking System</h1>"
            f"<p style='color:#93c5fd'>Daily Report — {datetime.datetime.utcnow().strftime('%d %B %Y')}</p>"
            "</div>"
            "<div style='padding:24px;border:1px solid #e5e7eb'>"
            f"{urgent_banner}"
            "<table style='width:100%;border-collapse:collapse'>"
            f"<tr><td style='padding:10px;border:1px solid #e5e7eb'>Total</td><td style='padding:10px;border:1px solid #e5e7eb'>{total}</td></tr>"
            f"<tr><td style='padding:10px;border:1px solid #e5e7eb'>Registered</td><td style='padding:10px;border:1px solid #e5e7eb'>{registered}</td></tr>"
            f"<tr><td style='padding:10px;border:1px solid #e5e7eb'>In Use</td><td style='padding:10px;border:1px solid #e5e7eb'>{in_use}</td></tr>"
            f"<tr><td style='padding:10px;border:1px solid #e5e7eb'>In Laundry</td><td style='padding:10px;border:1px solid #e5e7eb'>{in_laundry}</td></tr>"
            f"<tr><td style='padding:10px;border:1px solid #e5e7eb'>Missing</td><td style='padding:10px;border:1px solid #e5e7eb'>{missing}</td></tr>"
            "</table>"
            f"{missing_section}"
            f"{deleted_section}"
            "</div></div>"
        )

        send_email(subject, html_body)

# ============================================================
# STARTUP
# ============================================================
@asynccontextmanager
async def lifespan(app):
    init_db()
    asyncio.create_task(auto_mark_missing())
    asyncio.create_task(daily_report())
    print("Server started — background job running")
    yield
    print("Server shutting down")

app = FastAPI(lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============================================================
# DATA MODEL
# ============================================================
class TowelCreate(BaseModel):
    tag_id: str
    towel_type: str

# ============================================================
# ENDPOINTS
# ============================================================

@app.post("/towels")
def register_towel(towel: TowelCreate, _=Depends(verify_key)):
    db = SessionLocal()
    existing = db.query(Towel).filter(Towel.tag_id == towel.tag_id).first()
    if existing:
        raise HTTPException(status_code=400, detail="Tag ID already registered")
    new_towel = Towel(
        tag_id=towel.tag_id,
        towel_type=towel.towel_type,
        status="registered",
        last_location="Store",
        wash_count=0,
        created_at=datetime.datetime.utcnow()
    )
    db.add(new_towel)
    db.commit()
    event = Event(tag_id=towel.tag_id, event_type="REGISTERED", created_at=datetime.datetime.utcnow())
    db.add(event)
    db.commit()
    db.close()
    return {"message": "Towel registered successfully", "tag_id": towel.tag_id}


@app.get("/towels")
def get_all_towels():
    db = SessionLocal()
    towels = db.query(Towel).all()
    db.close()
    return towels


@app.get("/towels/{tag_id}")
def get_towel(tag_id: str):
    db = SessionLocal()
    towel = db.query(Towel).filter(Towel.tag_id == tag_id).first()
    db.close()
    if not towel:
        raise HTTPException(status_code=404, detail="Towel not found")
    return towel


@app.patch("/towels/{tag_id}/dispatch")
def dispatch_towel(tag_id: str, location: Optional[str] = None, _=Depends(verify_key)):
    db = SessionLocal()
    towel = db.query(Towel).filter(Towel.tag_id == tag_id).first()
    if not towel:
        raise HTTPException(status_code=404, detail="Towel not found")
    if towel.status == "in_use":
        raise HTTPException(status_code=400, detail="Towel already dispatched — return it first")
    if towel.wash_count >= MAX_WASHES:
        raise HTTPException(status_code=400, detail=f"Towel has exceeded {MAX_WASHES} wash cycles — retire and replace it")
    towel.status = "in_use"
    towel.dispatched_at = datetime.datetime.utcnow()
    towel.wash_count += 1
    if location:
        towel.last_location = location
    event = Event(tag_id=tag_id, event_type="DISPATCHED", location=location, created_at=datetime.datetime.utcnow())
    db.add(event)
    db.commit()
    result = {"message": "Towel dispatched", "tag_id": tag_id, "wash_count": towel.wash_count, "location": towel.last_location, "status": towel.status}
    db.close()
    return result


@app.patch("/towels/{tag_id}/return")
def return_towel(tag_id: str, _=Depends(verify_key)):
    db = SessionLocal()
    towel = db.query(Towel).filter(Towel.tag_id == tag_id).first()
    if not towel:
        raise HTTPException(status_code=404, detail="Towel not found")
    if towel.status not in ["in_use", "missing"]:
        raise HTTPException(status_code=400, detail=f"Towel is {towel.status} — only dispatched towels can be returned")
    towel.status = "in_laundry"
    towel.dispatched_at = None
    event = Event(tag_id=tag_id, event_type="RETURNED", created_at=datetime.datetime.utcnow())
    db.add(event)
    db.commit()
    result = {"message": "Towel returned to laundry", "tag_id": tag_id, "status": towel.status}
    db.close()
    return result


@app.patch("/towels/{tag_id}/clean")
def clean_towel(tag_id: str, _=Depends(verify_key)):
    db = SessionLocal()
    try:
        towel = db.query(Towel).filter(Towel.tag_id == tag_id).first()
        if not towel:
            raise HTTPException(status_code=404, detail="Towel not found")
        if towel.status != "in_laundry":
            raise HTTPException(status_code=400, detail=f"Towel is {towel.status} — only in_laundry towels can be marked clean")
        towel.status = "registered"
        event = Event(tag_id=tag_id, event_type="CLEANED", created_at=datetime.datetime.utcnow())
        db.add(event)
        db.commit()
        return {"message": "Towel marked as clean", "tag_id": tag_id, "status": towel.status, "wash_count": towel.wash_count}
    finally:
        db.close()


@app.patch("/towels/{tag_id}/retire")
def retire_towel(tag_id: str, reason: Optional[str] = None, _=Depends(verify_key)):
    db = SessionLocal()
    try:
        towel = db.query(Towel).filter(Towel.tag_id == tag_id).first()
        if not towel:
            raise HTTPException(status_code=404, detail="Towel not found")
        if towel.status == "in_use":
            raise HTTPException(status_code=400, detail="Towel is currently dispatched — return it first before retiring")
        if towel.status == "retired":
            raise HTTPException(status_code=400, detail="Towel is already retired")
        towel.status = "retired"
        event = Event(tag_id=tag_id, event_type="RETIRED", location=reason or "No reason provided", created_at=datetime.datetime.utcnow())
        db.add(event)
        db.commit()
        return {"message": "Towel retired successfully", "tag_id": tag_id, "status": towel.status, "total_washes": towel.wash_count, "reason": reason or "No reason provided"}
    finally:
        db.close()


@app.delete("/towels/{tag_id}")
def delete_towel(tag_id: str, reason: Optional[str] = None, _=Depends(verify_key)):
    db = SessionLocal()
    try:
        towel = db.query(Towel).filter(Towel.tag_id == tag_id).first()
        if not towel:
            raise HTTPException(status_code=404, detail="Towel not found")
        if towel.status == "in_use":
            raise HTTPException(status_code=400, detail="Towel is currently dispatched — return it first before deleting")
        deleted_log = DeletedTag(
            tag_id=towel.tag_id, towel_type=towel.towel_type, total_washes=towel.wash_count or 0,
            last_location=towel.last_location, reason=reason or "Defective RFID tag", deleted_at=datetime.datetime.utcnow()
        )
        db.add(deleted_log)
        db.query(Event).filter(Event.tag_id == tag_id).delete()
        db.delete(towel)
        db.commit()
        return {"message": f"Towel {tag_id} permanently deleted", "tag_id": tag_id, "total_washes": deleted_log.total_washes, "logged_to": "deleted_tags table"}
    finally:
        db.close()


@app.get("/deleted")
def get_deleted_tags(_=Depends(verify_key)):
    db = SessionLocal()
    try:
        deleted = db.query(DeletedTag).order_by(DeletedTag.deleted_at.desc()).all()
        return {
            "total_deleted": len(deleted),
            "tags": [{"tag_id": d.tag_id, "towel_type": d.towel_type, "total_washes": d.total_washes, "last_location": d.last_location, "reason": d.reason, "deleted_at": str(d.deleted_at)} for d in deleted]
        }
    finally:
        db.close()


@app.get("/missing")
def get_missing_towels():
    db = SessionLocal()
    threshold = datetime.datetime.utcnow() - datetime.timedelta(hours=24)
    missing = db.query(Towel).filter(
        (Towel.status == "missing") |
        ((Towel.status == "in_use") & (Towel.dispatched_at < threshold))
    ).all()
    db.close()
    return {"missing_count": len(missing), "towels": missing}


@app.get("/inventory")
def get_inventory():
    db = SessionLocal()
    try:
        total         = db.query(Towel).filter(Towel.status != "retired").count()
        registered    = db.query(Towel).filter(Towel.status == "registered").count()
        in_use        = db.query(Towel).filter(Towel.status == "in_use").count()
        in_laundry    = db.query(Towel).filter(Towel.status == "in_laundry").count()
        missing       = db.query(Towel).filter(Towel.status == "missing").count()
        retired       = db.query(Towel).filter(Towel.status == "retired").count()
        wash_exceeded = db.query(Towel).filter(Towel.wash_count >= MAX_WASHES, Towel.status != "retired").count()
        wash_critical = db.query(Towel).filter(Towel.wash_count >= WASH_CRITICAL, Towel.wash_count < MAX_WASHES, Towel.status != "retired").count()
        wash_warning  = db.query(Towel).filter(Towel.wash_count >= WASH_WARNING, Towel.wash_count < WASH_CRITICAL, Towel.status != "retired").count()
        return {"total": total, "registered": registered, "in_use": in_use, "in_laundry": in_laundry, "missing": missing, "retired": retired, "wash_exceeded": wash_exceeded, "wash_critical": wash_critical, "wash_warning": wash_warning}
    finally:
        db.close()


@app.get("/staff")
def staff_page(username: str = Depends(verify_staff)):
    from fastapi.responses import HTMLResponse
    base_dir = os.path.dirname(os.path.abspath(__file__))
    with open(os.path.join(base_dir, "staff.html"), "r") as f:
        return HTMLResponse(content=f.read())


@app.get("/dashboard")
def dashboard_page(username: str = Depends(verify_staff)):
    from fastapi.responses import HTMLResponse
    base_dir = os.path.dirname(os.path.abspath(__file__))
    with open(os.path.join(base_dir, "dashboard.html"), "r") as f:
        return HTMLResponse(content=f.read())


@app.get("/export")
def export_excel(status: Optional[str] = None, type: Optional[str] = None, location: Optional[str] = None, search: Optional[str] = None):
    db = SessionLocal()
    try:
        query = db.query(Towel)
        if status:   query = query.filter(Towel.status == status)
        if type:     query = query.filter(Towel.towel_type == type)
        if location: query = query.filter(Towel.last_location == location)
        if search:   query = query.filter(Towel.tag_id.ilike(f"%{search}%"))
        towels  = query.all()
        events  = db.query(Event).order_by(Event.created_at.desc()).all()
        missing = db.query(Towel).filter(Towel.status == "in_use", Towel.dispatched_at < datetime.datetime.utcnow() - datetime.timedelta(hours=24)).all()
        total_all  = db.query(Towel).count()
        registered = db.query(Towel).filter(Towel.status == "registered").count()
        in_use     = db.query(Towel).filter(Towel.status == "in_use").count()
        in_laundry = db.query(Towel).filter(Towel.status == "in_laundry").count()
        n_missing  = db.query(Towel).filter(Towel.status == "missing").count()

        HOTEL_BLUE = "1E3A5F"; WHITE = "FFFFFF"; HEADER_GRAY = "F8FAFC"; LIGHT_BLUE = "EFF6FF"
        GREEN_BG = "F0FDF4"; RED_BG = "FEF2F2"; ORANGE_BG = "FFFBEB"; BORDER_COLOR = "E2E8F0"
        STATUS_COLORS = {"in_use": "DBEAFE", "in_laundry": "FEF9C3", "registered": "DCFCE7", "missing": "FEE2E2"}
        EVENT_COLORS  = {"DISPATCHED": "DBEAFE", "RETURNED": "DCFCE7", "MISSING": "FEE2E2", "REGISTERED": "F3F4F6"}
        thin = Side(style='thin', color=BORDER_COLOR)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def hcell(ws, r, c, v, bg=HOTEL_BLUE, fg=WHITE):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = Font(name='Arial', bold=True, color=fg, size=11)
            cell.fill = PatternFill('solid', start_color=bg)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        def dcell(ws, r, c, v, bg=WHITE, bold=False, align='left'):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = Font(name='Arial', bold=bold, size=10)
            cell.fill = PatternFill('solid', start_color=bg)
            cell.alignment = Alignment(horizontal=align, vertical='center')
            cell.border = border

        def title_block(ws, title, subtitle):
            ws.row_dimensions[1].height = 36
            ws.merge_cells('A1:H1')
            c = ws['A1']; c.value = title
            c.font = Font(name='Arial', bold=True, size=16, color=WHITE)
            c.fill = PatternFill('solid', start_color=HOTEL_BLUE)
            c.alignment = Alignment(horizontal='left', vertical='center')
            ws.merge_cells('A2:H2')
            c2 = ws['A2']; c2.value = subtitle
            c2.font = Font(name='Arial', size=10, color="6B7280")
            c2.fill = PatternFill('solid', start_color=HEADER_GRAY)
            c2.alignment = Alignment(horizontal='left', vertical='center')

        generated_at = datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
        filter_desc = []
        if status:   filter_desc.append(f"Status: {status}")
        if type:     filter_desc.append(f"Type: {type}")
        if location: filter_desc.append(f"Location: {location}")
        if search:   filter_desc.append(f"Search: {search}")
        filter_note = "  |  Filters: " + ", ".join(filter_desc) if filter_desc else "  |  No filters — all data"

        wb = Workbook()
        wb.remove(wb.active)

        ws1 = wb.create_sheet("Summary")
        title_block(ws1, "Smart Linen Tracking — Inventory Summary", f"Generated: {generated_at}{filter_note}")
        for col, h in enumerate(["Metric", "Count", "% of Total", "Notes"], 1):
            hcell(ws1, 4, col, h)
        rows = [
            ("Total Towels", total_all,  None, "All towels in system", LIGHT_BLUE),
            ("Registered",   registered, registered/total_all  if total_all else 0, "In inventory, not dispatched", GREEN_BG),
            ("In Use",       in_use,     in_use/total_all       if total_all else 0, "Currently out in rooms", "DBEAFE"),
            ("In Laundry",   in_laundry, in_laundry/total_all   if total_all else 0, "Returned, being washed", ORANGE_BG),
            ("Missing",      n_missing,  n_missing/total_all    if total_all else 0, "Over 24hrs — investigate", RED_BG),
        ]
        for i, (metric, count, pct, note, bg) in enumerate(rows):
            r = i + 5
            dcell(ws1, r, 1, metric, bg=bg, bold=(i == 0))
            dcell(ws1, r, 2, count, bg=bg, align='center')
            if pct:
                cell = ws1.cell(row=r, column=3, value=pct)
                cell.number_format = '0.0%'
                cell.font = Font(name='Arial', size=10)
                cell.fill = PatternFill('solid', start_color=bg)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            else:
                dcell(ws1, r, 3, "—", bg=bg, align='center')
            dcell(ws1, r, 4, note, bg=bg)
        for col, w in enumerate([22, 12, 14, 36], 1):
            ws1.column_dimensions[get_column_letter(col)].width = w

        ws2 = wb.create_sheet("All Towels")
        title_block(ws2, f"Towels ({len(towels)} records)", f"Generated: {generated_at}{filter_note}")
        for col, h in enumerate(["Tag ID", "Type", "Status", "Last Location", "Wash Count", "Registered"], 1):
            hcell(ws2, 4, col, h)
        for i, t in enumerate(towels):
            r = i + 5; bg = STATUS_COLORS.get(t.status, WHITE)
            dcell(ws2, r, 1, t.tag_id, bg=bg)
            dcell(ws2, r, 2, (t.towel_type or "—").title(), bg=bg)
            dcell(ws2, r, 3, t.status.replace("_", " ").upper(), bg=bg, align='center')
            dcell(ws2, r, 4, t.last_location or "—", bg=bg)
            dcell(ws2, r, 5, t.wash_count or 0, bg=bg, align='center')
            dcell(ws2, r, 6, str(t.created_at or "—"), bg=bg)
        ws2.auto_filter.ref = f"A4:F{4+len(towels)}"
        ws2.freeze_panes = "A5"
        for col, w in enumerate([16, 14, 16, 18, 14, 20], 1):
            ws2.column_dimensions[get_column_letter(col)].width = w

        ws3 = wb.create_sheet("Missing Towels")
        title_block(ws3, "Missing Towels Report", f"Generated: {generated_at}")
        for col, h in enumerate(["Tag ID", "Type", "Last Location", "Dispatched At", "Action"], 1):
            hcell(ws3, 4, col, h, bg="991B1B")
        if missing:
            for i, t in enumerate(missing):
                r = i + 5
                dcell(ws3, r, 1, t.tag_id, bg=RED_BG)
                dcell(ws3, r, 2, (t.towel_type or "—").title(), bg=RED_BG)
                dcell(ws3, r, 3, t.last_location or "—", bg=RED_BG)
                dcell(ws3, r, 4, str(t.dispatched_at or "—"), bg=RED_BG)
                dcell(ws3, r, 5, "Investigate immediately", bg=RED_BG)
        else:
            ws3.merge_cells("A5:E5")
            c = ws3['A5']; c.value = "✓  No missing towels at time of export"
            c.font = Font(name='Arial', bold=True, size=11, color="166534")
            c.fill = PatternFill('solid', start_color=GREEN_BG)
            c.alignment = Alignment(horizontal='center', vertical='center')
            ws3.row_dimensions[5].height = 32
        ws3.auto_filter.ref = "A4:E4"
        ws3.freeze_panes = "A5"
        for col, w in enumerate([16, 14, 18, 22, 24], 1):
            ws3.column_dimensions[get_column_letter(col)].width = w

        ws4 = wb.create_sheet("Event History")
        title_block(ws4, "Event History Log", f"Generated: {generated_at}  |  {len(events)} events")
        for col, h in enumerate(["#", "Tag ID", "Event", "Location", "Date & Time"], 1):
            hcell(ws4, 4, col, h)
        for i, e in enumerate(events):
            r = i + 5; bg = EVENT_COLORS.get(e.event_type, WHITE)
            dcell(ws4, r, 1, i + 1, bg=bg, align='center')
            dcell(ws4, r, 2, e.tag_id, bg=bg)
            dcell(ws4, r, 3, e.event_type, bg=bg, align='center')
            dcell(ws4, r, 4, e.location or "—", bg=bg)
            dcell(ws4, r, 5, str(e.created_at or "—"), bg=bg)
        ws4.auto_filter.ref = f"A4:E{4+len(events)}"
        ws4.freeze_panes = "A5"
        for col, w in enumerate([6, 16, 16, 18, 24], 1):
            ws4.column_dimensions[get_column_letter(col)].width = w

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"linen_report_{datetime.datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    finally:
        db.close()
